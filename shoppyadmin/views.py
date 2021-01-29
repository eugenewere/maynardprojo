import csv
from datetime import datetime
from datetime import timedelta

from django.utils.crypto import get_random_string
from openpyxl import Workbook
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, HttpResponse, render_to_response

from django.template.loader import get_template
from django.views.generic.base import View
from openpyxl.styles.colors import COLOR_INDEX

from rest_framework.views import APIView
from rest_framework.response import Response
import datetime as dt
from shoppy.forms import *
from shoppy.models import *

from datetime import datetime
from django.utils import timezone
import pytz
from shoppyadmin.utils import render_to_pdf
from openpyxl.descriptors.excel import HexBinary, ExtensionList

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# from django.http import HttpResponse
@login_required
def home(request):
    user=request.user.id

    logged_in_seller= Seller.objects.filter(user_ptr_id=user).first()
    seller = Seller.objects.filter(id=request.user.id).first()

    products = Product.objects.filter(seller=logged_in_seller).order_by('-created_at')
    brand = Brand.objects.all()
    categories = Category.objects.all()


    sales = Order_payment.objects.filter(seller=seller)
    totalsales = []
    for sale in sales:
        # totalsales.append(sale.amount)
        totalsales.append((sale.amount) -- (sale.balance))
    sles = totalsales

    # .aggregate(sum=Sum('amount_paid'))['sum']
    # .annotate(sale=Count('amount_paid')).values('sale')
    # .annotate(sum=Sum('amount_paid'))['sum']
    # .aggregate(sum=Sum('amount_paid'))['sum']


    # checkouts = Checkout.objects.filter(status__exact='PENDING', order_product__product__seller=seller, order_product__checkout_id__isnull=False)
    checkouts = Checkout.objects.filter(status__exact='PENDING').order_by('-created_at')

    # buyers = []
    # for checkout in checkouts:
    #     buyers.append(checkout.buyer_with_this_checkout)
    # # print(buyers)
    # no_of_buyers_with_orders = list(set(buyers))


    prev_ords= Order_Product.objects.filter(product__seller_id=request.user.id, checkout_id__isnull=False, checkout__status='PAID')
    buyerz=[]
    order_buyer_history = Order_Product.objects.filter(product__seller_id=request.user.id).order_by('id')
    for orde in order_buyer_history:
        buyerz.append(orde.buyer)
    history_buyers_with_orders = list(set(buyerz))

    adminsales = Order_payment.objects.all()
    admintotalsales = []
    for sale in adminsales:
        admintotalsales.append(sale.amount)

    adminsles = list(set(admintotalsales))

    admin_orders = Checkout.objects.filter(status__exact='PENDING').order_by('-created_at')


    out_products = []
    inventories = Inventory.objects.filter(quantity__lt=1, product__seller=seller)
    for inventory in inventories:
        out_products.append(inventory.product)



    seller = Seller.objects.all()
    context = {
        'ord_count':checkouts.count(),
        'seller': seller,
        'prev_ords':prev_ords,
        'admin_prev_ords':Order_Product.objects.filter( checkout_id__isnull=False, checkout__status='PAID'),
        'out_of_stock_products':out_products,

        'products' : products,
        'brands' : brand,
        'categories' : categories,
        'admin_buyers' :Buyer.objects.all(),
        'admin_sellers' :Seller.objects.all(),
        'buyer_orders': checkouts,
        # 'buyer_orders': list(set(buyers)),
        'admin_buyer_orders': admin_orders,
        'sales':sum(i for i in sles),
        'admin_sales':sum(i for i in adminsles),
        'history_buyers_with_orders':history_buyers_with_orders,
        # 'order_buyers':checkouts

    }
    return render(request ,'index.html' ,context)
# def trial():
#     orders = Order_Product.objects.filter(product__seller=request.user.id, checkout_id__isnull=False)
#     buyers = []
#     for order in orders:
#         buyers.append(order.buyer)
# product section
@login_required()
def view_all_products(request):
    user = request.user.id
    seller = Seller.objects.filter(user_ptr_id=user).first()
    if request.method == 'POST' and request.is_ajax():
        name = request.POST.get('name')
        bran = request.POST.get('product_brand')
        sdesc = request.POST.get('short_description')
        ldesc = request.POST.get('long_description')
        img = request.FILES.get('featured_url')
        vat = request.POST.get('vat_status')
        categor = request.POST.get('category')
        invtry=request.POST.get('quantity')
        brand = Brand.objects.filter(id=bran).first()
        category = Category.objects.filter(id=categor).first()
        images = request.FILES.getlist('other_images[]')
        cost = request.POST.get('unit_cost')
        form = ProductForm(request.POST, request.FILES)
        print(form)

        if form.is_valid():
            new_product = Product.objects.create(
                name=name,
                unit_cost=cost,
                product_brand=brand,
                short_description=sdesc,
                long_description=ldesc,
                seller=seller,
                featured_url=img,
                vat_status=vat,
                category=category,
            )
            Inventory.objects.create(
                product=new_product,
                quantity=invtry,
            )
            for upload in images:
                image_file = {
                    'image': upload,
                }
                image_form = ImagesForm({'product': new_product.id}, image_file)
                image_form.save()

            for variant_option_id in request.POST.getlist('variant_options[]'):
                variant_option = Variant_Option.objects.filter(id=int(variant_option_id)).first()

                if variant_option is not None:
                    Product_Variant_Options.objects.create(
                        product=new_product,
                        variant_options=variant_option
                    )
            data = {
                'results': 'success',
                'success': 'Product Added Successfully added'
            }
            return JsonResponse(data, safe=False)
        else:
            form1 = ProductForm(request.POST, request.FILES)
            data = {
                'results': 'error',
                'error': 'Error Adding The Product',
                'form': form1,
            }
            return render_to_response('products/error.html', data)
    return redirect('ShoppyAdmin:view_products')

def view_products(request):
    if request.method == "POST":
        seller_id = request.POST['seller']
        productz = Product.objects.filter(seller_id=seller_id).order_by('status')
    else:
        productz = Product.objects.order_by('status')



    brands = Brand.objects.all()
    categories = Category.objects.all()
    user = request.user.id
    seller = Seller.objects.filter(user_ptr_id=user).first()
    products = Product.objects.filter(seller=seller).order_by('-created_at')
    variants = Variant.objects.filter(seller_id=seller.id)

    context = {
        'seller': Seller.objects.all(),
        'products': products,
        'sellerz':seller,
        'adminProducts':productz,
        'brands': brands,
        'categories': categories,
        # 'variantOptions':variantOptions,
        'variants': variants,
    }

    return render(request, 'products/view_products.html', context)



@login_required()
def edit_product(request, product_id):
    product = Product.objects.filter(id=product_id).first()
    variants = Variant.objects.all()
    images = product.images
    if request.method == 'POST':
        product_form = ProductForm(request.POST, request.FILES , instance=product)
        # print(product_form)
        if product_form.is_valid():
           new_product = product_form.save()

           for upload in request.FILES.getlist('other_images[]'):
               print(upload)
               Image.objects.filter(product=new_product).update_or_create(
                   product=new_product,
                   image = upload,
               )

           messages.success(request, 'Product Edited Successfully')
        else:
            messages.error(request, 'Error editing')
    context = {
        'product': product,
        'images': images,
        'variants': variants,
        'categories': Category.objects.all(),
        'brands':Brand.objects.all(),

    }

    return render(request, 'products/edit_products.html', context)

@login_required()
def product_delete(request, product_id):
    product = Product.objects.filter(id=product_id)
    product.delete()
    messages.success(request, 'Product Deleted Successfully')

    return redirect('ShoppyAdmin:shoppy_admin_view_all_products')

@login_required()
def image_delete(request, image_id):
    image = Image.objects.filter(id=image_id).first()
    if image is not None:
        product_id = image.product_id
        image.delete()
        messages.success(request, 'Image deleted successfully')

        return redirect('ShoppyAdmin:shoppy_admin_edit_product', product_id)


# brand section
@login_required()
def add_brand(request):
    if request.method == 'POST':
        brand_form = AddBrandForm(request.POST)

        if brand_form.is_valid():
            brand_form.save()
            messages.success(request, 'Brand Added Successfully')
            return redirect('ShoppyAdmin:shoppy_admin_add_brand')
        else:
            messages.error(request, 'Brand Not Added')
            return redirect('ShoppyAdmin:shoppy_admin_add_brand')

    brands = Brand.objects.all()
    categories = Category.objects.all()

    context = {
        'brands': brands,
        'categories': categories,
    }
    return render(request, 'brand and category/brand.html', context)

@login_required()
def brand_edit(request, brand_id):
    if request.method == 'POST':
        brand = Brand.objects.get(id=brand_id)
        form = AddBrandForm(request.POST, instance=brand)
        if form.is_valid():
            form.save()
            messages.success(request, 'Brand Update Successful')
            return redirect('ShoppyAdmin:shoppy_admin_add_brand')
        else:
            messages.error(request, 'Form Invalid')
            return redirect('ShoppyAdmin:shoppy_admin_add_brand')

@login_required()
def brand_delete(request, brand_id):
    brand = Brand.objects.filter(id=brand_id).first()
    brand.delete()
    messages.success(request, 'Brand Deleted Successfully')
    return redirect('ShoppyAdmin:shoppy_admin_add_brand')


# category section... this is to view category
@login_required()
def view_category(request):
    independent_categories = []
    categories = Category.objects.all()
    for category in categories:
        if not category.is_child:
            independent_categories.append(category)

    context = {
        'categories': set(independent_categories),
    }

    return render(request, 'brand and category/category.html', context)

@login_required()
def view_sub_category(request, category_id):
    sub_categories = Category.objects.filter(parent_id=category_id)
    categories = Category.objects.all()
    context = {
        "categories": categories,
        "sub_categories": sub_categories,
        # 'brand_count':new_brand_count,

    }

    return render(request, "brand and category/sub_categories.html", context)

@login_required()
def addingcategory(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        category = request.POST['name']
        if not Category.objects.filter(name__exact=category):
            if form.is_valid():
                form.save()
                messages.success(request, 'Category Added Successfully')
            else:
                messages.error(request, 'Category Not Added')
        else:
            messages.error(request ,'Category Exists')
    return redirect('ShoppyAdmin:shoppy_admin_add_category')

@login_required()
def deletecategory(request, category_id):
    category = Category.objects.filter(id=category_id)
    category.delete()
    return redirect("ShoppyAdmin:shoppy_admin_add_category")

@login_required()
def editcategory(request, category_id):
    category = Category.objects.filter(id=category_id).first()
    categoryform = CategoryForm(request.POST, instance=category)
    if categoryform.is_valid():
        categoryform.save()
        messages.success(request, "Category updated successfully")
    else:
        messages.error(request, "Error Updating Category")
    return redirect("ShoppyAdmin:shoppy_admin_add_category")


# end of categories

# seller
@login_required()
def sellers(request):
    seller = Seller.objects.all()

    context = {
        'seller': seller,
    }
    return render(request, 'sellers.html', context)

@login_required()
def sellerStatusChange(request, seller_id):
    seller = Seller.objects.filter(id=seller_id).first()

    if seller.status == "UNVERIFIED":
        Seller.objects.filter(id=seller.id).update(
            status='VERIFIED'
        )
        Product.objects.filter(seller=seller.id).update(
            status = 'VERIFIED'
        )
        messages.success(request, 'Seller Is Activated')
    else:
        Seller.objects.filter(id=seller.id).update(
            status='UNVERIFIED'
        )
        Product.objects.filter(seller=seller.id).update(
            status='UNVERIFIED'
        )
        messages.success(request, 'Seller Is Deactivated')
    return redirect('ShoppyAdmin:shoppy-admin-sellers')


# seller

@login_required()
def orders(request):
    user = request.user.id
    independant_orders = []

    if request.method == "POST":
        seller_id = request.POST['seller']

        orderz = Order_Product.objects.filter(product__seller=seller_id, checkout__isnull=False, checkout__status='PENDING').order_by("-created_at")
    else:
        orderz = Order_Product.objects.filter(checkout__isnull=False, checkout__status='PENDING').order_by("-created_at")

    seller = Seller.objects.filter(user_ptr_id=user).first()
    orders = Order_Product.objects.filter(product__seller=seller, checkout__isnull=False, checkout__status='PENDING').order_by("-created_at")



    context = {
        'buyer_orders':Checkout.objects.order_by('-created_at'),
        'adminorders':orderz,
        'seller':Seller.objects.all()

    }
    return render(request, 'orders.html', context)

@login_required()
def user_login(request):
    return render(request, 'login.html')

@login_required()
def reviews(request):
    comments = []
    user = request.user.id
    seller = Seller.objects.filter(user_ptr_id=user).first()
    reviews = Review.objects.filter(product__seller=user).first()
    products = Product.objects.filter(seller=seller)

    for product in products:
        review = Review.objects.filter(product=product).first()
        if review is not None:
            comments.append(review)
    context={
        'seller':seller,
        'reviews': set(comments)
    }
    return render(request, 'reviews.html', context)

@login_required()
def single_product_review(request, product_id):
    user = request.user.id
    reviews = Review.objects.filter(product__seller=user, product=product_id)
    proreview=Review.objects.filter(product=product_id).first()
    admin_reviews = Review.objects.filter( product=product_id)
    admin_proreview=Review.objects.filter(product=product_id).first()

    # product = Product.objects.filter(id=review.product.id)
    context = {
        'reviews': reviews ,
        'proreview':proreview,
        'admin_reviews': admin_reviews ,
        'admin_proreview':admin_proreview,
    }

    return render(request, 'singleproductreviews.html', context)

def edit_payments(request, Order_payment_id):
    orderPay = Order_payment.objects.filter(id=Order_payment_id).first()
    dbcheckout = Checkout.objects.filter(order_payment=orderPay).first()

    if request.method == "POST":

        dbamount = orderPay.amount
        dbpayrefcode =  orderPay.payment_refrence
        dbbalance =orderPay.balance
        dbshipping = dbcheckout.shipping_cost

        amount = request.POST["amount"]
        # ref_code = request.POST["ref_code"]
        payment_reference = request.POST["payment_reference"]
        shippingcost = request.POST["shipping_cost"]


        if shippingcost is not None:
            shipping_cost = shippingcost
        else:
            shipping_cost = dbshipping



        if dbpayrefcode == payment_reference:
            newUpdatedpayref = payment_reference
        else:
            newUpdatedpayref = dbpayrefcode

        form_checkout = Checkout.objects.filter(order_payment = orderPay).first()

        UpdatedShipping = float(shipping_cost)
        newUpdatedAmmount = float(dbamount) + float(amount)
        balance = float(dbbalance) - float(amount)
        new_amount = float(UpdatedShipping) + float(newUpdatedAmmount)
        newpayref = newUpdatedpayref
        negativebalance = float(form_checkout.total) - float(newUpdatedAmmount)


        if float(newUpdatedAmmount) >= float(form_checkout.total) :
            Order_payment.objects.filter(id=orderPay.id).update(
                checkout_id=form_checkout.id,
                amount=float(new_amount),
                payment_refrence=newpayref,
                balance= float(negativebalance),
            )
            checkout=Checkout.objects.filter(id=form_checkout.id).update(
                status='PAID',
                shipping_cost=float(UpdatedShipping),
                amount_paid=float(new_amount),
            )
            orders = Order_Product.objects.filter(checkout=form_checkout.id)
            for order in orders:
                product_inventory_count = order.product.inventory_qty
                order_quantity = order.quantity
                remaining_quantity = product_inventory_count - order_quantity
                print(product_inventory_count, order_quantity, remaining_quantity)
                Inventory.objects.filter(product_id=order.product.id).update(
                    quantity=remaining_quantity,
                )

            messages.success(request, 'Order Payment Is Completed')
        else:
            Order_payment.objects.filter(id=orderPay.id).update(
                checkout_id=form_checkout.id,
                amount=float(new_amount),
                payment_refrence=newpayref,
                balance=float(balance),
            )
            Checkout.objects.filter(id=form_checkout.id).update(
                status='PENDING',
                amount_paid=float(new_amount),
                shipping_cost = float(UpdatedShipping)
            )
            messages.info(request, 'Order Payment Is Not Payed Fully')
    return redirect('ShoppyAdmin:shoppy-payments')


def payments(request):
    seller = Seller.objects.filter(user_ptr_id=request.user.id).first()
    if request.method == "POST":

        amount = request.POST["amount"]
        ref_code = request.POST["ref_code"]
        payment_reference = request.POST["payment_reference"]
        shipping_cost = request.POST["shipping_cost"]
        form_checkout = Checkout.objects.filter(reference_code__iexact=ref_code).first()

        if shipping_cost is not None:
            new_shipping_cost =shipping_cost
        else:
            new_shipping_cost = 0

        balance = float(form_checkout.total) - float(amount)
        new_amount = float(new_shipping_cost) + float(amount)



        if float(amount) >= float(form_checkout.total):
            Order_payment.objects.create(
                checkout_id=form_checkout.id,
                amount=float(new_amount),
                seller=seller,
                payment_refrence=payment_reference,
                balance=float(balance),
            )
            Checkout.objects.filter(id=form_checkout.id).update(
                status='PAID',
                shipping_cost=float(new_shipping_cost),
                amount_paid=amount,
            )


            orders = Order_Product.objects.filter(checkout=form_checkout.id)

            for order in orders:
                print(order)
                product_inventory_count =  order.product.inventory_qty
                order_quantity = order.quantity
                remaining_quantity = (product_inventory_count) - (order_quantity)
                print(product_inventory_count, order_quantity, remaining_quantity)
                Inventory.objects.filter(product=order.product).update(
                    quantity=int(remaining_quantity),
                )
            messages.success(request, 'Order Payment Is Completed')
        else:
            Order_payment.objects.create(
                checkout_id=form_checkout.id,
                amount=float(amount),
                seller=seller,
                payment_refrence=payment_reference,
                balance=float(balance),

            )
            Checkout.objects.filter(id=form_checkout.id).update(
                status='PENDING',
                amount_paid=amount,
                shipping_cost = float(new_shipping_cost)
            )
            messages.info(request, 'Order Payment Is Not Payed Fully')

    paymentsz=[]
    sellerrr = Seller.objects.filter(id=request.user.id).first()
    payments = Order_payment.objects.filter(seller=sellerrr).order_by('-created_at')
    for pay in payments:
        paymentsz.append(pay)

    checkoutsz = []
    checkouts = Checkout.objects.filter(order_payment__payment_refrence__isnull=True, status='PENDING', order_product__product__seller_id=request.user.id,)
    for checkout in checkouts:
        checkoutsz.append(checkout)
    context={
        # 'checkouts':checkoutsz,
        'checkouts':list(set(checkoutsz)),
        'payments':paymentsz
        # 'payments':list(set(paymentsz)),
    }
    return render(request, 'payments/payments.html', context)














@login_required()
def buyers(request):
    buyers = Buyer.objects.all()

    context = {
        'buyers': buyers,
    }
    return render(request, 'buyers.html', context)


@login_required()
def userAccount(request):
    user = request.user
    seller = Seller.objects.filter(user_ptr_id=user.id).first()

    if request.method == 'POST':
        seller_form = SellerForm(request.POST, request.FILES, instance=seller)

        seller_form.save()

        username = request.POST['username']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']

        if seller is not None:
            user = User.objects.filter(pk=request.user.id).update(
                first_name=first_name,
                last_name=last_name,
                username=username
            )
            # user.first_name = first_name,
            # user.last_name = last_name,
            # user.username = username,
            # user.save()

        messages.success(request, 'Account Updated Successfully')
        return redirect("ShoppyAdmin:shoppy-user-account")

    context = {
        'seller': seller,
    }
    return render(request, 'useraccount.html', context)

@login_required()
def adminAccount(request):
    user = request.user
    admin = User.objects.filter(id = user.id).first()

    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']

        if admin is not None:
            user = User.objects.filter(id=admin.id).update(
                first_name=first_name,
                last_name=last_name,
                username=username,
                email=email,
            )
            # user.first_name = first_name,
            # user.last_name = last_name,
            # user.username = username,
            # user.save()

        messages.success(request, 'Account Updated Successfully')
        return redirect("ShoppyAdmin:shoppy-user-account")


    return render(request, 'useraccount.html')

def changepassword(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('ShoppyAdmin:shoppy-changepassword')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request,'accounts/changepassword.html',{'form':form})

# carousel
@login_required()
def view_carousel(request):
    if request.method == 'POST':
        carouselForm = CarouselImageForm(request.POST, request.FILES)
        if carouselForm.is_valid():
            carouselForm.save()
            messages.success(request, 'Advertisment Image Added Successfully')
            return redirect('ShoppyAdmin:view_carousel')
        else:
            messages.error(request, 'Advertisment Image not Added ')
            return redirect('ShoppyAdmin:view_carousel')

    else:
        carousels = Carousel.objects.all()
        carouselForm = CarouselImageForm()


    context = {
        'carousels': carousels,
        'form': carouselForm,
        'seller':Seller.objects.filter(status='VERIFIED'),
        'products':Product.objects.filter(status='VERIFIED')

    }

    return render(request, 'products/carousel.html', context)


def carousel_delete(request, carousel_id):
    carousel = Carousel.objects.filter(id=carousel_id).first()
    carousel.delete()
    messages.success(request, 'Advertisment Image Deleted Succesfully')

    return redirect('ShoppyAdmin:view_carousel')

def carousel_edit(request, carousel_id):
    carousel = Carousel.objects.filter(id=carousel_id).first()
    form = CarouselImageForm(request.POST, request.FILES, instance=carousel)
    if form.is_valid():
        form.save()
        messages.success(request, 'Advertisment Image Updated Succesfully')

    return redirect('ShoppyAdmin:view_carousel')


# end_of_carousel


# list,edit and delete regions
@login_required()
def view_regions(request):
    user=request.user.id
    seller =Seller.objects.filter(id=user).first()
    if request.method == 'POST':
        region=request.POST['name']
        region_cost = request.POST['region_cost']

        Region.objects.create(
            name=region,
            region_cost=region_cost,
            seller=seller,
        )
        messages.success(request, 'A Region And Its Respective Cost Has Been Added')
        return redirect("ShoppyAdmin:shoppy-admin-view-regions")



    regions = Region.objects.filter(seller=seller)
    sellers = Seller.objects.all()

    context = {
        'regions': regions,
        'sellers': sellers,
    }
    return render(request, 'regions/regions.html', context)

@login_required()
def edit_regions(request, region_id):
    if request.method == 'POST':
        region = request.POST['name']
        region_cost = request.POST['region_cost']
        Region.objects.filter(id=region_id).update(
            name=region,
            region_cost=region_cost,
        )
        messages.success(request, 'Region Update Successful')
        return redirect('ShoppyAdmin:shoppy-admin-view-regions')


@login_required()
def delete_regions(request, region_id):
    region = Region.objects.filter(id=region_id).first()
    region.delete()
    messages.success(request, 'Region Deleted Successfully')
    return redirect('ShoppyAdmin:shoppy-admin-view-regions')


# end_of_regions


# variant,variantoptions ( add,delete,edit)
@login_required()
def variants(request):
    user=request.user.id
    seller = Seller.objects.filter(user_ptr_id=user).first()
    if request.method == 'POST':

        variant_form = VariantForm(request.POST)
        if variant_form.is_valid():
            variant_form.save()
            messages.success(request, 'Variant Added Successfully')
            return redirect('ShoppyAdmin:variants')
        else:
            messages.error(request, 'Error Adding Variants')
            return redirect('ShoppyAdmin:variants')
    variants = Variant.objects.filter(seller=seller).all()
    variant_options = Variant_Option.objects.filter(seller=seller).order_by("variant")
    context = {
        'variants': variants,
        'variant_options': variant_options,
        'seller':seller,
    }
    return render(request, 'variants/variants.html', context)

@login_required()
def variant_delete(request, variant_id):
    variant = Variant.objects.filter(id=variant_id).first()
    variant.delete()
    messages.success(request, 'Variant Deleted Successfully')
    return redirect('ShoppyAdmin:variants')


def variant_edit(request, variant_id):
    if request.method == 'POST':
        variant = Variant.objects.filter(id=variant_id).first()
        variant_form = VariantForm(request.POST, instance=variant)
        if variant_form.is_valid():
            variant_form.save()
            messages.success(request, 'Variant Update Successful')
            return redirect('ShoppyAdmin:variants')
        else:
            messages.success(request, 'Variant Error Updating')
            return redirect('ShoppyAdmin:variants')

@login_required()
def variants_options(request):
    user = request.user.id
    seller = Seller.objects.filter(user_ptr_id=user).first()

    if request.method == 'POST':
        variantOptions = VariantOptionForm(request.POST)
        if variantOptions.is_valid():
            variantOptions.save()
            messages.success(request, 'Variant Options Added Successfully')
            return redirect('ShoppyAdmin:variants')
        else:
            messages.error(request, 'Variant Option Not Added')
    return redirect('ShoppyAdmin:variants')

@login_required()
def variants_options_edit(request, variant_option_id):
    if request.method == 'POST':
        variant_option = Variant_Option.objects.filter(id=variant_option_id).first()
        variant_option_form = VariantOptionForm(request.POST, instance=variant_option)
        if variant_option_form.is_valid():
            variant_option_form.save()
            messages.success(request, 'Variant Options Update Successful')
            return redirect('ShoppyAdmin:variants')
        else:
            messages.success(request, 'Variant Options Error Updating')
            return redirect('ShoppyAdmin:variants')

@login_required()
def variants_options_delete(request, variant_option_id):
    variantOption = Variant_Option.objects.filter(id=variant_option_id).first()
    variantOption.delete()
    messages.success(request, 'Variant Option Deleted Successfully')
    return redirect('ShoppyAdmin:variants')


# end_of_variant


# featuredproducts
def featured_products(request):
    user = request.user.id
    featured_products = Product.objects.filter(feat_product="FEATURED", seller=user)
    admin_featured_products = Product.objects.filter(feat_product='FEATURED')
    context = {
        'products' : featured_products,
        'f_product' : admin_featured_products,
    }

    return render(request, 'products/featuredproducts.html', context)

@login_required()
def featured_products_featured(request, product_id):
    product = Product.objects.filter(id=product_id).first()
    if product.feat_product == "FEATURED":
        Product.objects.filter(id=product_id).update(feat_product="NORMAL")
        messages.success(request, 'Product Status Is Updated to Normal Successfully')
    return redirect('ShoppyAdmin:shoppy_admin_view_all_products')

@login_required()
def normal_products_normal(request, product_id):
    product = Product.objects.filter(id=product_id).first()
    if product.feat_product == "NORMAL":
        Product.objects.filter(id=product_id).update(feat_product="FEATURED")
        messages.success(request, 'Product Status Is Updated to Featured Successfully')
    return redirect('ShoppyAdmin:shoppy_admin_view_all_products')

@login_required()
def verify_product(request, product_id):
    product = Product.objects.filter(id=product_id).first()
    if product.status == "UNVERIFIED":
        Product.objects.filter(id=product_id).update(status="VERIFIED")
        messages.success(request, 'Product Status Is Verified')
    return redirect('ShoppyAdmin:shoppy_admin_view_all_products')

@csrf_exempt
@login_required()
def product_featured(request):
    if request.is_ajax():
        product_id = request.POST.get('product_id')
        product = Product.objects.filter(id=product_id).first()
        if product.feat_product == "FEATURED":
            Product.objects.filter(id=product_id).update(feat_product="NORMAL")
            context = {
                'results': 'normal_success'
            }
        elif product.feat_product == "NORMAL":
            Product.objects.filter(id=product_id).update(feat_product="FEATURED")
            context = {
                'results': 'featured_success'
            }
        return JsonResponse(context)











@login_required()
def unverify_product(request, product_id):
    product = Product.objects.filter(id=product_id).first()
    if product.status == "VERIFIED":
        Product.objects.filter(id=product_id).update(status="UNVERIFIED")
        messages.success(request, 'Product Status Is UnVerified')
    return redirect('ShoppyAdmin:shoppy_admin_view_all_products')

@csrf_exempt
def product_verification(request):
    if request.is_ajax() :
        product_id = request.POST.get('product_id')
        product = Product.objects.filter(id=product_id).first()
        if product.status == "VERIFIED":
            Product.objects.filter(id=product_id).update(status="UNVERIFIED")
            context = {
                'results':'unverified_success'
            }
        elif product.status == "UNVERIFIED":
            Product.objects.filter(id=product_id).update(status="VERIFIED")
            context = {
                'results': 'verified_success'
            }
        return JsonResponse(context)

# end of featured products
# offer
@login_required()
def offer(request):
    user=request.user.id
    products = Product.objects.filter(seller=user)
    offers=Offer.objects.filter(product__seller=user).order_by('-created_at')
    all_offers = Offer.objects.all()

    context = {
        'products': products,
        'offers': offers,
        'all_offers': all_offers,
    }
    return render(request, 'products/productoffers.html', context)
@login_required()
def addOffer(request):
    if request.method == 'POST':
        form = OfferForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Offer Successfully Been Made')
            return redirect('ShoppyAdmin:view_product_offer')
        messages.error(request, 'Error Adding An Offer')
        return redirect('ShoppyAdmin:view_product_offer')

    return redirect('ShoppyAdmin:view_product_offer')
@login_required()
def editOffer(request,offer_id):
    offer =Offer.objects.filter(id=offer_id).first()
    if offer:
        form = OfferForm(request.POST, instance=offer)
        if request.method == 'POST':
            if form.is_valid():
               form.save()
               messages.success(request, "Offer Updated Successfully")
               return redirect('ShoppyAdmin:view_product_offer')
            else:
                messages.error(request, "Offer Not Updated Successfully")
                return redirect('ShoppyAdmin:view_product_offer')
    else:
        messages.error(request,"Offer Not Found")
        return redirect('ShoppyAdmin:view_product_offer')
@login_required()
def deleteOffer(request, offer_id):
    offer =Offer.objects.filter(id=offer_id).first()

    if offer:

        offer.delete()
        messages.success(request,'Offer Deleted Successfuly')
        return redirect('ShoppyAdmin:view_product_offer')
    else:
        messages.error(request, 'Error! Offer Not Found')
        return redirect('ShoppyAdmin:view_product_offer')



# offer
@login_required()
def viewAllReports(request):

    return render(request,'reports/reports.html')
# reports

class GeneratePDF(View):
    def get(self, request, *args, **kwargs):
        orders = Order_Product.objects.filter(checkout__isnull=False)
        template = get_template('invoice.html')
        context = {
            'orders': orders,
        }
        html = template.render(context)
        pdf = render_to_pdf('invoice.html', context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Invoice_%s.pdf" % ("12341231")
            content = "inline; filename='%s'" % (filename)
            # download = request.GET.get("download")
            response['Content-Disposition'] = content
            # if download:
            #     content = "attachment; filename='%s'" % (filename)
            return pdf
        return HttpResponse("Not found")

class GenerateProductPDF(View):
    def get(self, request, *args, **kwargs):
        user = request.user.id
        products =Product.objects.filter(seller=user)
        template = get_template('productspdf.html')
        context = {
            'products': products,
        }
        html = template.render(context)
        pdf = render_to_pdf('productspdf.html', context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Invoice_%s.pdf" % ("12341231")
            content = "inline; filename='%s'" % (filename)
            # download = request.GET.get("download")
            response['Content-Disposition'] = content
            # if download:
            #     content = "attachment; filename='%s'" % (filename)
            return pdf
        return HttpResponse("Not found")

# reports

class ListUsers(APIView):
    authentication_classes = []
    permission_classes = []
    def get(self, request, format=None):
        # from datetime import datetime
        import datetime
        sellers=Seller.objects.all().count()
        buyers=Buyer.objects.all().count()
        user = request.user.id
        seller = Seller.objects.filter(user_ptr_id=user).first()
        month_data = []
        seller_count = []
        buyer_count = []
        months_choices = []
        months_choices_int = []
        for i in range(1, 13):
            months_choices.append((datetime.date(2008, i, 1).strftime('%B')[0:3]))
        labels2 = months_choices
        for z in range(1, 13):
            months_choices_int.append((datetime.date(2008, z, 1).strftime('%m')))
        for months_choice in months_choices_int:

            # month_data.append(Order_Product.objects.filter(checkout__isnull=False, created_at__month=months_choice, product__seller=seller).count())
            seller_count.append(Seller.objects.filter(date_joined__month=months_choice).count())
        defaultData2 =list(seller_count)
        for months_choice in months_choices_int:

            # month_data.append(Order_Product.objects.filter(checkout__isnull=False, created_at__month=months_choice, product__seller=seller).count())
            buyer_count.append(Buyer.objects.filter(date_joined__month=months_choice).count())
        defaultData = buyer_count



        labels = ['Sellers','Buyers']

        context={
            'labels':labels2,
            'defaultData2':defaultData2,
            'defaultData':defaultData,

        }
        return Response(context)


def ListOrders(request):
    user=request.user.id
    import datetime
    month_data = []
    months_choices=[]
    months_choices_int=[]
    for i in range(1,13):
        months_choices.append(( datetime.date(2008, i, 1).strftime('%B')[0:3]))

    labels2 = months_choices

    for z in range(1,13):
        months_choices_int.append((datetime.date(2008, z, 1).strftime('%m')))

    seller = Seller.objects.filter(id=user).first()
    for months_choice in months_choices_int:
        month_data.append(Order_Product.objects.filter(checkout__status='PAID', product__seller=seller,  created_at__month=months_choice).count())
    defaultData2 = month_data
    context2={
        'labels2':labels2,
        'defaultData23':defaultData2,


    }

    month_data = {}
    return JsonResponse(context2)

class ListOrdersadmin(APIView):
    authentication_classes = []
    permission_classes = []
    def get(self, request, format=None):
        user=request.user.id
        import datetime
        month_data = []
        months_choices=[]
        months_choices_int=[]
        for i in range(1,13):
            months_choices.append(( datetime.date(2008, i, 1).strftime('%B')[0:3]))

        labels2 = months_choices

        for z in range(1,13):
            months_choices_int.append((datetime.date(2008, z, 1).strftime('%m')))

        for months_choice in months_choices_int:
            month_data.append(Order_Product.objects.filter(checkout__isnull=False, created_at__month=months_choice).count())

        defaultData2 = month_data
        context2={
            'labels3':labels2,
            'defaultData3':defaultData2,


        }

        month_data = {}
        return Response(context2)



def buyerOrders(request, reference_code):

    buyer_orders = []
    admin_orders = []
    user = request.user.id


    seller = Seller.objects.filter(user_ptr_id=user).first()

    orders = Order_Product.objects.filter(
        checkout__reference_code__exact=reference_code,
        product__seller=seller,
    ).order_by('-created_at')
    # print(orders)
    # for order in orders:
    #     buyer_orders.append(order)

    orderss = Order_Product.objects.filter(checkout__reference_code__exact=reference_code, checkout__isnull=False, checkout__status__iexact='PENDING',).order_by('-created_at')



    context = {
        # 'buyer':buyer,
        'orders':orders,
        'admin_orders':orderss
    }

    return render(request, 'buyer_orders/buyer_order.html', context)

def buyerOrdersCancel(request,reference_code):
    checkout = Checkout.objects.filter(reference_code__exact=reference_code).first()
    print(checkout)
    orders = Order_Product.objects.filter(checkout=checkout)
    print(orders)
    if checkout is not None:
        for order in orders:
            order.delete()
        checkout.delete()
        messages.success(request, "Order Deleted Successfully")
    return redirect('ShoppyAdmin:shoppy-admin-home')


def export_users_xls(request):
    """
        Downloads all movies as Excel file with a single worksheet
    """
    if request.method=='POST':
        user = request.POST.get('buyer', False)
        datefrom = request.POST['datefrom']
        dateto = request.POST['dateto']
        import datetime
        buyer = Buyer.objects.filter(user_ptr_id=user).first()
        order_queryset = Order_Product.objects.filter(product__seller_id=request.user.id, checkout__status='PAID', buyer=buyer, created_at__range=[datefrom, dateto])
    else:
        order_queryset = Order_Product.objects.filter(product__seller_id=request.user.id, checkout__status='PAID')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    from datetime import datetime
    response['Content-Disposition'] = 'attachment; filename={date}-orders.xlsx'.format(

        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # # Get active worksheet/tab
    # worksheet = workbook.active
    # worksheet.title = 'Orders'

    # Define the titles for columns
    columns = [
        ('Product', 30),
        # ('Variants', 15 ),
        ('Buyer', 20),
        ('Phone_Number', 15),
        ('Quantity', 15),
        ('Receipt_No',15),
        ('Order_Number', 15),
        ('Shipping_Cost', 15),
        ('Order_total', 15),
        ('Amount_Paid', 15),
        ('Order_Date', 15),
    ]
    # row_num = 1
    # Iterate through movie categories
    for category_index, category in enumerate(order_queryset):
        # Create a worksheet/tab with the title of the category
        worksheet = workbook.create_sheet(
            title=category.product.name,
            index=category_index,
        )
        # Define the background color of the header cells
        fill = PatternFill(
            start_color="00ffb53e",
            end_color="00243144",
            fill_type='solid',

        )
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.fill = fill
            # set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

        # Iterate through all movies
        for movie in order_queryset:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                (movie.product.name, 'Normal'),
                # (movie.order_variants_for_excel, 'Normal'),
                (movie.buyer.first_name + ' '+ movie.buyer.last_name, 'Normal'),
                (movie.checkout.phonenumber, 'Normal'),
                (movie.quantity, 'Normal'),
                (movie.orderpaymentttt.payment_refrence, 'Normal' ),
                (movie.checkout.reference_code, 'Normal'),
                (movie.checkout.shipping_cost, 'Normal'),
                (movie.total , 'Normal'),
                (movie.checkout.amount_paid, 'Normal'),
                (datetime.strftime(movie.created_at, "%d/%b/%Y"), 'Normal',),
            ]


            # Assign values, styles, and formatting for each cell in the row
            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                if cell_format == 'Currency':
                    cell.number_format = '#,##0 €'
                # if col_num == 4:
                #     cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment
        # freeze the first row
        worksheet.freeze_panes = worksheet['A2']
        # set tab color
        worksheet.sheet_properties.tabColor = COLOR_INDEX[5]

        workbook.save(response)

    return response


def export_users_xls_date(request):
    if request.method=='POST':

        datefrom = request.POST['datefrom']
        dateto = request.POST['dateto']
        import datetime
        order_queryset = Order_Product.objects.filter(product__seller_id=request.user.id, checkout__status='PAID', created_at__range=[datefrom, dateto])
    else:
        order_queryset = Order_Product.objects.filter(product__seller_id=request.user.id, checkout__status='PAID')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    from datetime import datetime
    response['Content-Disposition'] = 'attachment; filename={date}-orders.xlsx'.format(

        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # # Get active worksheet/tab
    # worksheet = workbook.active
    # worksheet.title = 'Orders'

    # Define the titles for columns
    columns = [
        ('Product', 30),
        # ('Variants', 15 ),
        ('Buyer', 10),
        ('Phone_Number', 15),
        ('Quantity', 15),
        ('Order_Number', 15),
        ('Shipping_Cost', 15),
        ('Order_total', 15),
        ('Amount_Paid', 15),
        ('Order_Date', 15),
    ]
    # row_num = 1
    # Iterate through movie categories
    for category_index, category in enumerate(order_queryset):
        # Create a worksheet/tab with the title of the category
        worksheet = workbook.create_sheet(
            title=category.product.name,
            index=category_index,
        )
        # Define the background color of the header cells
        fill = PatternFill(
            start_color="00ffb53e",
            end_color="00243144",
            fill_type='solid',

        )
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border_bottom
            cell.alignment = centered_alignment
            cell.fill = fill
            # set column width
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

        # Iterate through all movies
        for movie in order_queryset:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                (movie.product.name, 'Normal'),
                # (movie.order_variants_for_excel, 'Normal'),
                (movie.buyer.first_name + ' '+ movie.buyer.last_name, 'Normal'),
                (movie.checkout.phonenumber, 'Normal'),
                (movie.quantity, 'Normal'),
                (movie.checkout.reference_code, 'Normal'),
                (movie.checkout.shipping_cost, 'Normal'),
                (movie.total, 'Normal'),
                (movie.checkout.amount_paid, 'Normal'),
                (datetime.strftime(movie.created_at, "%d/%b/%Y"), 'Normal',),
            ]

            # Assign values, styles, and formatting for each cell in the row
            for col_num, (cell_value, cell_format) in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.style = cell_format
                if cell_format == 'Currency':
                    cell.number_format = '#,##0.00 €'
                # if col_num == 4:
                #     cell.number_format = '[h]:mm;@'
                cell.alignment = wrapped_alignment
        # freeze the first row
        worksheet.freeze_panes = worksheet['A2']
        # set tab color
        worksheet.sheet_properties.tabColor = COLOR_INDEX[5]

        workbook.save(response)

    return response


def update_Inventory(request, product_id):
    product = Product.objects.filter(id=product_id).first()
    inventory = Inventory.objects.filter(product=product).first()
    if request.method== 'POST':
        inventory_input = request.POST['inventory']
        if not int(inventory_input) <= 0:
            Inventory.objects.filter(id=inventory.id).update(
                quantity=inventory_input
            )
            messages.success(request, "Product Inventory Updated Successfully")
        else:
            messages.error(request, "Please update with appropriate quantity")
    return redirect('ShoppyAdmin:shoppy-admin-home')


def voucher(request):
    vouchers = Voucher.objects.order_by('-created_at')
    context = {
        'vouchers':vouchers,
    }
    return render(request, 'voucher/voucher.html', context)


def addvoucher(request):
    if request.method == 'POST':
        event = request.POST['event']
        amount = request.POST['amount']
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        no_of_users = request.POST['no_of_users']
        code = get_random_string(5, 'ABCDEFGHIJKLMNPQRSTUVWXYZ123456789')
        seller=Seller.objects.filter(user_ptr_id=request.user.id).first()
        Voucher.objects.create(
            code=code,
            event=event,
            seller=seller,
            amount=amount,
            start_time=start_date,
            end_time=end_date,
            no_of_users=no_of_users,
        )
        messages.success(request, 'Vocher added successfully')
    else:
        messages.error(request, 'Error adding a voucher')

    return redirect('ShoppyAdmin:voucher')


def deletevoucher(request, voucher_id):
    voucher = Voucher.objects.filter(id=voucher_id).first()
    print(voucher, voucher_id)

    if voucher:
        voucher.delete()
        messages.success(request, 'Voucher Deleted Successfuly')
        return redirect('ShoppyAdmin:voucher')
    else:
        messages.error(request, 'Error! Voucher Not Found')
        return redirect('ShoppyAdmin:voucher')


def editvoucher(request, voucher_id):
    voucher = Voucher.objects.filter(id=voucher_id).first()
    if voucher and request.method == 'POST':
        event = request.POST['event']
        amount = request.POST['amount']
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        no_of_users = request.POST['no_of_users']
        seller = Seller.objects.filter(user_ptr_id=request.user.id).first()
        Voucher.objects.filter(id=voucher.id).update(
            event=event,
            seller=seller,
            amount=amount,
            start_time=start_date,
            end_time=end_date,
            no_of_users=no_of_users,
        )
        messages.success(request, "Vocher Updated Successfully")
        return redirect('ShoppyAdmin:voucher')

    else:
        messages.error(request, "Vocher Not Found")
        return redirect('ShoppyAdmin:voucher')